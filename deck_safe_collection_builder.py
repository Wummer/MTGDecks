#!/usr/bin/env python3
"""
Deck-Safe Collection Builder
Builds a deck-safe collection spreadsheet from a Moxfield CSV export and deck .txt files.

Usage:
    python deck_safe_collection_builder.py <collection.csv> <deck1.txt> [deck2.txt ...] [-o output.xlsx]

Or point at a directory of .txt files:
    python deck_safe_collection_builder.py <collection.csv> --deck-dir ./decks/ [-o output.xlsx]

The script will:
- Parse your Moxfield collection CSV (proxies count as owned)
- Parse each deck .txt file (99 main + 1 commander, sideboard = maybeboard)
- Resolve card names across DFCs, split cards, and UB reskin aliases
- Produce a multi-tab Excel spreadsheet with allocation analysis
"""

import csv
import os
import re
import sys
import argparse
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ============================================================
# CONFIGURATION — Update aliases here as needed
# ============================================================
ALIASES = {
    "Aang's Shelter": "Teferi's Protection",
    "The Banyan Tree": "The Great Henge",
    "Lifelong Friendship": "Eladamri's Call",
    "Castle Shimura": "Eiganjo Castle",
    "Wild Rose Rebellion": "Counterspell",
    "Paradise Chocobo": "Birds of Paradise",
    "Joo Dee, Public Servant": "Sakashima of a Thousand Faces",
    "Expansion // Explosion": "Expansion/Explosion",
    "Lady Octopus, Inspired Inventor": "Merata, Neuron Hacker",
    "Find // Finality": "Find/Finality",
    "Dawn Warriors' Legacy": "Mizzix's Mastery",
    "Bayo, Irritable Instructor":"Electro, Assaulting Battery",

}
REVERSE_ALIASES = {v: k for k, v in ALIASES.items()}


def canonical_name(card_name):
    return ALIASES.get(card_name, card_name)


# ============================================================
# COLLECTION PARSING
# ============================================================
def parse_collection(csv_path):
    collection_raw = defaultdict(int)
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            name = row['Name'].strip()
            count = int(row['Count'].strip())
            collection_raw[name] += count

    collection = defaultdict(int)
    for name, count in collection_raw.items():
        collection[name] += count
        if '//' in name:
            front = name.split('//')[0].strip()
            collection[front] += count
        cn = canonical_name(name)
        if cn != name:
            collection[cn] += count

    return collection_raw, collection


def resolve_owned(card_name, collection):
    cn = canonical_name(card_name)
    names = {cn, card_name}
    if cn in REVERSE_ALIASES:
        names.add(REVERSE_ALIASES[cn])
    if card_name in ALIASES:
        names.add(ALIASES[card_name])
    return max(collection.get(n, 0) for n in names)


# ============================================================
# DECK PARSING
# ============================================================
def parse_deck(filepath):
    lines = open(filepath, encoding='utf-8').read().strip().split('\n')

    sb_idx = None
    blank_after_sb = None
    for i, line in enumerate(lines):
        if line.strip() == 'SIDEBOARD:':
            sb_idx = i
        elif sb_idx is not None and line.strip() == '' and blank_after_sb is None:
            blank_after_sb = i

    main_cards, sb_cards, commander = {}, {}, None

    if sb_idx is None:
        last_blank = None
        for i, line in enumerate(lines):
            if line.strip() == '':
                last_blank = i
        for i, line in enumerate(lines):
            ls = line.strip()
            if not ls:
                continue
            m = re.match(r'^(\d+)\s+(.+)$', ls)
            if m:
                cnt, card = int(m.group(1)), m.group(2).strip()
                if last_blank is not None and i > last_blank:
                    commander = card
                else:
                    main_cards[card] = main_cards.get(card, 0) + cnt
    else:
        for i, line in enumerate(lines):
            ls = line.strip()
            if not ls or ls == 'SIDEBOARD:':
                continue
            m = re.match(r'^(\d+)\s+(.+)$', ls)
            if m:
                cnt, card = int(m.group(1)), m.group(2).strip()
                if i < sb_idx:
                    main_cards[card] = main_cards.get(card, 0) + cnt
                elif blank_after_sb is not None and i > blank_after_sb:
                    commander = card
                else:
                    sb_cards[card] = sb_cards.get(card, 0) + cnt

    actual_deck = dict(main_cards)
    if commander:
        actual_deck[commander] = actual_deck.get(commander, 0) + 1

    return actual_deck, sb_cards, commander


def clean_deck_name(filename):
    n = filename.replace('.txt', '')
    n = re.sub(r'-\d{8}-\d{6}$', '', n)
    n = re.sub(r'-\d{8}$', '', n)
    for s in ['-duplicated-from-tumultuoustempus', '-duplicated-from-sleepyheaded',
              '--duplicated-from-kaboomeow', '-updated']:
        n = n.replace(s, '')
    return n.replace('-', ' ').title()


# ============================================================
# GOOGLE SHEETS ADAPTER
# ============================================================
class _ColumnDim:
    def __init__(self):
        self.width = None

class _ColumnDims:
    def __init__(self):
        self._dims = {}
    def __getitem__(self, key):
        if key not in self._dims:
            self._dims[key] = _ColumnDim()
        return self._dims[key]

class _SheetProps:
    def __init__(self):
        self.tabColor = None

def _col_letter_to_num(letter):
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result

def _parse_cell_ref(ref):
    m = re.match(r'^([A-Z]+)(\d+)$', ref.upper())
    if not m:
        raise ValueError(f"Invalid cell reference: {ref}")
    return int(m.group(2)), _col_letter_to_num(m.group(1))


class GCell:
    def __init__(self, row, col, value=None):
        self.row = row
        self.col = col
        self.value = value
        self.font = None
        self.fill = None
        self.alignment = None
        self.number_format = None


class GWorksheet:
    def __init__(self, title):
        self.title = title
        self.cells = {}
        self.column_dimensions = _ColumnDims()
        self._freeze_panes = None
        self._merges = []
        self.sheet_properties = _SheetProps()

    def cell(self, row, column, value=None):
        key = (row, column)
        if key not in self.cells:
            self.cells[key] = GCell(row, column)
        c = self.cells[key]
        if value is not None:
            c.value = value
        return c

    def __getitem__(self, ref):
        row, col = _parse_cell_ref(ref)
        return self.cell(row, col)

    def __setitem__(self, ref, value):
        row, col = _parse_cell_ref(ref)
        self.cell(row, col, value)

    @property
    def freeze_panes(self):
        return self._freeze_panes

    @freeze_panes.setter
    def freeze_panes(self, value):
        self._freeze_panes = value

    def merge_cells(self, range_string):
        self._merges.append(range_string)


class GWorkbook:
    def __init__(self):
        self._sheets = []
        self._active = GWorksheet("Sheet1")
        self._sheets.append(self._active)

    @property
    def active(self):
        return self._active

    def create_sheet(self, title):
        ws = GWorksheet(title)
        self._sheets.append(ws)
        return ws

    def flush_to_google(self, spreadsheet):
        from gspread.utils import a1_range_to_grid_range
        from gspread_formatting import CellFormat, Color, TextFormat, NumberFormat
        import gspread_formatting.batch_update_requests as bur

        def hex_to_color(h):
            h = h.lstrip('#')
            if len(h) == 6:
                return Color(int(h[0:2], 16) / 255, int(h[2:4], 16) / 255, int(h[4:6], 16) / 255)
            return Color(0, 0, 0)

        def translate_font(font):
            if font is None:
                return {}
            kwargs = {}
            if font.bold:
                kwargs['bold'] = True
            if font.italic:
                kwargs['italic'] = True
            if font.color and font.color.rgb and font.color.rgb != '00000000':
                rgb = font.color.rgb
                if len(rgb) == 8:
                    rgb = rgb[2:]  # strip alpha
                kwargs['foregroundColor'] = hex_to_color(rgb)
            if font.name:
                kwargs['fontFamily'] = font.name
            if font.size:
                kwargs['fontSize'] = font.size
            if kwargs:
                return {'textFormat': TextFormat(**kwargs)}
            return {}

        def translate_fill(fill):
            if fill is None or fill.fill_type is None:
                return {}
            fg = fill.fgColor
            if fg and fg.rgb and fg.rgb != '00000000':
                rgb = fg.rgb
                if len(rgb) == 8:
                    rgb = rgb[2:]
                return {'backgroundColor': hex_to_color(rgb)}
            return {}

        def translate_alignment(alignment):
            if alignment is None:
                return {}
            kwargs = {}
            if alignment.horizontal:
                kwargs['horizontalAlignment'] = alignment.horizontal.upper()
            if hasattr(alignment, 'text_rotation') and alignment.text_rotation:
                kwargs['textRotation'] = {'angle': alignment.text_rotation}
            return kwargs

        def translate_number_format(nf):
            if nf is None or nf == 'General':
                return {}
            nf_type = 'NUMBER'
            if '%' in nf:
                nf_type = 'PERCENT'
            return {'numberFormat': NumberFormat(type=nf_type, pattern=nf)}

        # ── Phase 1: Sheet structure ──────────────────────────────────────────
        existing = spreadsheet.worksheets()
        existing_titles = {ws.title for ws in existing}

        for i, gws in enumerate(self._sheets):
            if i == 0:
                if existing[0].title != gws.title:
                    existing[0].update_title(gws.title)
            elif gws.title not in existing_titles:
                spreadsheet.add_worksheet(title=gws.title, rows=1000, cols=26)

        our_titles = {gws.title for gws in self._sheets}
        refreshed = spreadsheet.worksheets()
        for ws in refreshed:
            if ws.title not in our_titles and len(refreshed) > 1:
                spreadsheet.del_worksheet(ws)

        ws_cache = {ws.title: ws for ws in spreadsheet.worksheets()}

        # ── Phase 2: Clear all sheets in one call ─────────────────────────────
        sheets_with_data = [gws for gws in self._sheets if gws.cells]
        if sheets_with_data:
            spreadsheet.values_batch_clear(
                body={'ranges': [gws.title for gws in sheets_with_data]}
            )

        # ── Phase 3: Write all values in one call ─────────────────────────────
        value_ranges = []
        sheet_bounds = {}  # title → (max_row, max_col)
        for gws in sheets_with_data:
            max_row = max(r for r, c in gws.cells.keys())
            max_col = max(c for r, c in gws.cells.keys())
            sheet_bounds[gws.title] = (max_row, max_col)
            values = [
                [
                    (gws.cells[(r, c)].value if (r, c) in gws.cells and gws.cells[(r, c)].value is not None else '')
                    for c in range(1, max_col + 1)
                ]
                for r in range(1, max_row + 1)
            ]
            value_ranges.append({
                'range': f"'{gws.title}'!A1:{get_column_letter(max_col)}{max_row}",
                'values': values,
            })

        if value_ranges:
            spreadsheet.values_batch_update(body={
                'data': value_ranges,
                'valueInputOption': 'RAW',
            })

        # ── Phase 4: All formatting in one batch_update ───────────────────────
        all_requests = []
        for gws in sheets_with_data:
            ws = ws_cache[gws.title]
            max_row, max_col = sheet_bounds[gws.title]

            # Resize if needed
            if max_row > ws.row_count or max_col > ws.col_count:
                all_requests.append({
                    'updateSheetProperties': {
                        'properties': {
                            'sheetId': ws.id,
                            'gridProperties': {
                                'rowCount': max(max_row, ws.row_count),
                                'columnCount': max(max_col, ws.col_count),
                            },
                        },
                        'fields': 'gridProperties.rowCount,gridProperties.columnCount',
                    }
                })

            # Cell formatting
            fmt_ranges = []
            for (r, c), cell in gws.cells.items():
                fmt_kwargs = {}
                if cell.font:
                    fmt_kwargs.update(translate_font(cell.font))
                if cell.fill:
                    fmt_kwargs.update(translate_fill(cell.fill))
                if cell.alignment:
                    fmt_kwargs.update(translate_alignment(cell.alignment))
                if cell.number_format:
                    fmt_kwargs.update(translate_number_format(cell.number_format))
                if fmt_kwargs:
                    fmt_ranges.append((f'{get_column_letter(c)}{r}', CellFormat(**fmt_kwargs)))
            if fmt_ranges:
                all_requests.extend(bur.format_cell_ranges(ws, fmt_ranges))

            # Merges
            for merge_range in gws._merges:
                all_requests.append({
                    'mergeCells': {
                        'mergeType': 'MERGE_ALL',
                        'range': a1_range_to_grid_range(merge_range, ws.id),
                    }
                })

            # Frozen panes
            if gws._freeze_panes:
                fp_row, fp_col = _parse_cell_ref(gws._freeze_panes)
                freeze_kwargs = {}
                if fp_row > 1:
                    freeze_kwargs['rows'] = fp_row - 1
                if fp_col > 1:
                    freeze_kwargs['cols'] = fp_col - 1
                if freeze_kwargs:
                    all_requests.extend(bur.set_frozen(ws, **freeze_kwargs))

            # Column widths (all columns for this sheet in one extend)
            col_widths = [
                (col, int(dim.width * 7))
                for col, dim in gws.column_dimensions._dims.items()
                if dim.width
            ]
            if col_widths:
                all_requests.extend(bur.set_column_widths(ws, col_widths))

            # Tab color
            if gws.sheet_properties.tabColor:
                color = hex_to_color(gws.sheet_properties.tabColor)
                all_requests.append({
                    'updateSheetProperties': {
                        'properties': {
                            'sheetId': ws.id,
                            'tabColor': {
                                'red': color.red or 0,
                                'green': color.green or 0,
                                'blue': color.blue or 0,
                            },
                        },
                        'fields': 'tabColor',
                    }
                })

        if all_requests:
            spreadsheet.batch_update({'requests': all_requests})


# ============================================================
# CONSIDERING DECKS ANALYSIS
# ============================================================
def compute_available_pool(collection, resolved_demand):
    """Return {canonical_name: available_count} after existing decks claim their cards."""
    available = {}
    # Cards demanded by existing decks: available = max(0, surplus)
    for cn, info in resolved_demand.items():
        available[cn] = max(0, info['surplus'])
    # Cards in collection not demanded by any existing deck: fully available
    for card, count in collection.items():
        cn = canonical_name(card)
        if cn not in available:
            available[cn] = max(available.get(cn, 0), count)
    return available


def compute_assembly_order(considering_decks, considering_commanders,
                           available_pool, collection):
    """Greedy allocation: pick highest-% completable deck, claim its cards, repeat."""
    pool = dict(available_pool)
    remaining = set(considering_decks.keys())
    result = []

    while remaining:
        candidates = []
        for deck_name in remaining:
            cards = considering_decks[deck_name]
            total = sum(cards.values())
            have = 0
            missing_cards = {}
            for card, needed in cards.items():
                cn = canonical_name(card)
                avail = pool.get(cn, 0)
                # Also check collection directly for cards not tracked in pool yet
                if cn not in pool:
                    avail = resolve_owned(card, collection)
                    pool[cn] = avail
                got = min(needed, avail)
                have += got
                if got < needed:
                    missing_cards[card] = needed - got
            pct = have / total if total > 0 else 0
            candidates.append((pct, total - have, deck_name, have, total, missing_cards))

        # Best: highest pct, then fewest missing, then alphabetical
        candidates.sort(key=lambda x: (-x[0], x[1], x[2]))
        _, _, deck_name, have, total, missing_cards = candidates[0]

        # Claim cards from pool
        for card, needed in considering_decks[deck_name].items():
            cn = canonical_name(card)
            claimed = min(needed, pool.get(cn, 0))
            pool[cn] = pool.get(cn, 0) - claimed

        result.append({
            'name': deck_name,
            'order': len(result) + 1,
            'total': total,
            'have': have,
            'missing': total - have,
            'pct': pct,
            'missing_cards': missing_cards,
            'commander': considering_commanders.get(deck_name, ''),
        })
        remaining.remove(deck_name)

    return result


def write_proxy_files(assembly_order, output_dir):
    """Write one proxy .txt per considering deck with missing cards only."""
    os.makedirs(output_dir, exist_ok=True)
    written = []
    for entry in assembly_order:
        missing = entry['missing_cards']
        if not missing:
            continue
        safe_name = clean_deck_name(entry['name'] + '.txt').lower().replace(' ', '-')
        filepath = os.path.join(output_dir, f"{safe_name}-proxy.txt")
        with open(filepath, 'w', encoding='utf-8') as f:
            for card in sorted(missing.keys()):
                f.write(f"{missing[card]} {card}\n")
        written.append(filepath)
    return written


# ============================================================
# SPREADSHEET BUILDING
# ============================================================
def build_spreadsheet(collection_raw, collection, decks, maybeboard, commanders,
                      deck_names_raw, deck_display, resolved_demand, resolve_cache,
                      csv_basename, output_path=None, workbook=None,
                      considering_decks=None, considering_commanders=None,
                      assembly_order=None, available_pool=None):
    wb = workbook or Workbook()
    HF = PatternFill('solid', fgColor='4472C4')
    HN = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    NF = Font(name='Arial', size=10)
    BF = Font(bold=True, name='Arial', size=10)
    RF = Font(name='Arial', size=10, color='FF0000', bold=True)
    GF = Font(name='Arial', size=10, color='008000')
    RED = PatternFill('solid', fgColor='FFCCCC')
    YEL = PatternFill('solid', fgColor='FFFFCC')
    GRN = PatternFill('solid', fgColor='CCFFCC')
    CTR = Alignment(horizontal='center')

    missing = sum(1 for v in resolved_demand.values() if v['surplus'] < 0)
    shared = sum(1 for v in resolved_demand.values() if len(v['decks']) > 1)

    def hdr(ws, row, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HN; c.fill = HF; c.alignment = CTR

    # --- SUMMARY ---
    ws = wb.active; ws.title = "Summary"; ws.sheet_properties.tabColor = "4472C4"
    ws['A1'] = "Deck-Safe Collection Report"
    ws['A1'].font = Font(bold=True, name='Arial', size=16, color='4472C4')
    ws.merge_cells('A1:D1')
    ws['A3'] = f"Generated from {csv_basename} across {len(decks)} Commander decks"
    ws['A3'].font = Font(name='Arial', size=10, color='666666', italic=True)
    ws['A4'] = "Proxies counted as owned. DFCs matched by front face. Considering/maybeboard excluded."
    ws['A4'].font = Font(name='Arial', size=9, color='666666', italic=True)
    ws['A5'] = f"Reskins: {', '.join(f'{k} = {v}' for k, v in ALIASES.items())}"
    ws['A5'].font = Font(name='Arial', size=9, color='666666', italic=True)

    row = 7
    for label, val in [
        ("Total Unique Cards in Collection", len(collection_raw)),
        ("Total Copies in Collection", sum(collection_raw.values())),
        ("", ""),
        ("Unique Cards Needed (Main + Commander)", len(resolved_demand)),
        ("Total Copies Needed", sum(v['total_demand'] for v in resolved_demand.values())),
        ("", ""),
        ("Cards Shared Between 2+ Decks", shared),
        ("Cards with Insufficient Copies", missing),
        ("Total Additional Copies Needed", abs(sum(v['surplus'] for v in resolved_demand.values() if v['surplus'] < 0))),
        ("Cards Fully Covered", sum(1 for v in resolved_demand.values() if v['surplus'] >= 0)),
    ]:
        ws.cell(row=row, column=1, value=label).font = BF if label else NF
        if val != "":
            c = ws.cell(row=row, column=3, value=val); c.font = NF; c.alignment = Alignment(horizontal='right')
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Deck Breakdown").font = Font(bold=True, name='Arial', size=13, color='4472C4')
    row += 1
    hdr(ws, row, ["Deck", "Commander", "Cards", "Owned", "Missing", "% Complete"])
    row += 1
    for dn in deck_names_raw:
        cards = decks[dn]
        total = sum(cards.values())
        owned_count = sum(min(cnt, resolve_cache.get(card, [card, 0])[1]) for card, cnt in cards.items())
        miss = total - owned_count
        pct = owned_count / total if total > 0 else 0
        ws.cell(row=row, column=1, value=deck_display[dn]).font = NF
        ws.cell(row=row, column=2, value=commanders.get(dn, "")).font = Font(name='Arial', size=10, italic=True)
        for ci, val in [(3, total), (4, owned_count)]:
            ws.cell(row=row, column=ci, value=val).font = NF; ws.cell(row=row, column=ci).alignment = CTR
        c = ws.cell(row=row, column=5, value=miss); c.font = RF if miss > 0 else GF; c.alignment = CTR
        c = ws.cell(row=row, column=6, value=pct); c.number_format = '0.0%'; c.font = NF; c.alignment = CTR
        c.fill = GRN if pct >= 1.0 else (YEL if pct >= 0.9 else RED)
        row += 1
    ws.column_dimensions['A'].width = 42; ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10; ws.column_dimensions['F'].width = 14

    # Considering decks mini-summary on the Summary tab
    if assembly_order:
        row += 1
        ws.cell(row=row, column=1, value="Considering Decks (Available After Existing Decks)").font = Font(bold=True, name='Arial', size=13, color='5B9BD5')
        row += 1
        hdr(ws, row, ["Deck", "Commander", "Cards", "Available", "Missing", "% Complete"])
        row += 1
        for entry in assembly_order:
            display = clean_deck_name(entry['name'] + '.txt')
            ws.cell(row=row, column=1, value=display).font = NF
            ws.cell(row=row, column=2, value=entry['commander']).font = Font(name='Arial', size=10, italic=True)
            for ci, val in [(3, entry['total']), (4, entry['have'])]:
                ws.cell(row=row, column=ci, value=val).font = NF; ws.cell(row=row, column=ci).alignment = CTR
            c = ws.cell(row=row, column=5, value=entry['missing']); c.font = RF if entry['missing'] > 0 else GF; c.alignment = CTR
            c = ws.cell(row=row, column=6, value=entry['pct']); c.number_format = '0.0%'; c.font = NF; c.alignment = CTR
            c.fill = GRN if entry['pct'] >= 1.0 else (YEL if entry['pct'] >= 0.9 else RED)
            row += 1
        ws.cell(row=row, column=1, value="See 'Assembly Order' tab for recommended build sequence").font = Font(name='Arial', size=9, italic=True, color='888888')

    # --- SHOPPING LIST ---
    ws2 = wb.create_sheet("Shopping List"); ws2.sheet_properties.tabColor = "FF0000"
    hdr(ws2, 1, ["Card Name", "Also Known As", "Copies Needed", "Total Demand", "Owned", "Needed By"])
    row = 2
    for card, info in sorted(((k, v) for k, v in resolved_demand.items() if v['surplus'] < 0), key=lambda x: x[1]['surplus']):
        aka = ", ".join(n for n in info.get('display_names', []) if n != card)
        ws2.cell(row=row, column=1, value=card).font = NF
        ws2.cell(row=row, column=2, value=aka).font = Font(name='Arial', size=9, color='888888', italic=True)
        c = ws2.cell(row=row, column=3, value=abs(info['surplus'])); c.font = RF; c.alignment = CTR
        ws2.cell(row=row, column=4, value=info['total_demand']).font = NF; ws2.cell(row=row, column=4).alignment = CTR
        ws2.cell(row=row, column=5, value=info['owned']).font = NF; ws2.cell(row=row, column=5).alignment = CTR
        ws2.cell(row=row, column=6, value=", ".join(deck_display[d] for d in info['decks'])).font = NF
        row += 1
    ws2.column_dimensions['A'].width = 42; ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 16; ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 10; ws2.column_dimensions['F'].width = 80
    ws2.freeze_panes = 'A2'

    # --- SHARED CARDS ---
    ws3 = wb.create_sheet("Shared Cards"); ws3.sheet_properties.tabColor = "FFC000"
    hdr(ws3, 1, ["Card Name", "Also Known As", "# Decks", "Total Demand", "Owned", "Surplus/Deficit", "Decks"])
    row = 2
    for card, info in sorted(((k, v) for k, v in resolved_demand.items() if len(v['decks']) > 1), key=lambda x: (-len(x[1]['decks']), x[1]['surplus'])):
        aka = ", ".join(n for n in info.get('display_names', []) if n != card)
        ws3.cell(row=row, column=1, value=card).font = NF
        ws3.cell(row=row, column=2, value=aka).font = Font(name='Arial', size=9, color='888888', italic=True)
        ws3.cell(row=row, column=3, value=len(info['decks'])).font = BF; ws3.cell(row=row, column=3).alignment = CTR
        ws3.cell(row=row, column=4, value=info['total_demand']).font = NF; ws3.cell(row=row, column=4).alignment = CTR
        ws3.cell(row=row, column=5, value=info['owned']).font = NF; ws3.cell(row=row, column=5).alignment = CTR
        s = info['surplus']
        c = ws3.cell(row=row, column=6, value=s); c.alignment = CTR
        c.font = RF if s < 0 else (BF if s == 0 else GF)
        c.fill = RED if s < 0 else (YEL if s == 0 else GRN)
        ws3.cell(row=row, column=7, value=", ".join(deck_display[d] for d in info['decks'])).font = NF
        row += 1
    ws3.column_dimensions['A'].width = 42; ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 10; ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 10; ws3.column_dimensions['F'].width = 16; ws3.column_dimensions['G'].width = 80
    ws3.freeze_panes = 'A2'

    # --- FULL MATRIX ---
    ws4 = wb.create_sheet("Full Card Matrix"); ws4.sheet_properties.tabColor = "70AD47"
    hdr(ws4, 1, ["Card Name", "Owned", "Total Demand", "Surplus"])
    for di, dn in enumerate(deck_names_raw):
        c = ws4.cell(row=1, column=5 + di, value=deck_display[dn])
        c.font = HN; c.fill = HF; c.alignment = Alignment(horizontal='center', text_rotation=45)
    row = 2
    for card, info in sorted(resolved_demand.items(), key=lambda x: (x[1]['surplus'], -len(x[1]['decks']), x[0])):
        display_names = info.get('display_names', [card])
        label = card + (f" ({'/'.join(n for n in display_names if n != card)})" if any(n != card for n in display_names) else "")
        ws4.cell(row=row, column=1, value=label).font = NF
        ws4.cell(row=row, column=2, value=info['owned']).font = NF; ws4.cell(row=row, column=2).alignment = CTR
        ws4.cell(row=row, column=3, value=info['total_demand']).font = NF; ws4.cell(row=row, column=3).alignment = CTR
        s = info['surplus']
        c = ws4.cell(row=row, column=4, value=s); c.alignment = CTR
        c.font = RF if s < 0 else (BF if s == 0 else GF)
        c.fill = RED if s < 0 else (YEL if s == 0 else GRN)
        for di, dn in enumerate(deck_names_raw):
            cnt = sum(dcnt for dc, dcnt in decks[dn].items() if canonical_name(dc) == card)
            if cnt > 0:
                c = ws4.cell(row=row, column=5 + di, value=cnt); c.font = NF; c.alignment = CTR
                if s < 0: c.fill = PatternFill('solid', fgColor='FFF2CC')
        row += 1
    ws4.column_dimensions['A'].width = 50; ws4.column_dimensions['B'].width = 10
    ws4.column_dimensions['C'].width = 14; ws4.column_dimensions['D'].width = 12
    for di in range(len(deck_names_raw)):
        ws4.column_dimensions[get_column_letter(5 + di)].width = 18
    ws4.freeze_panes = 'E2'

    # --- RESKIN ALIASES ---
    ws5 = wb.create_sheet("Reskin Aliases"); ws5.sheet_properties.tabColor = "9966CC"
    hdr(ws5, 1, ["UB / Reskin Name", "Original MTG Name", "Used In", "Owned", "Status"])
    row = 2
    for ub_name, orig_name in sorted(ALIASES.items()):
        cn = canonical_name(ub_name)
        info = resolved_demand.get(cn, {'owned': 0, 'surplus': 0, 'decks': []})
        ws5.cell(row=row, column=1, value=ub_name).font = NF
        ws5.cell(row=row, column=2, value=orig_name).font = NF
        ws5.cell(row=row, column=3, value=", ".join(deck_display[d] for d in info.get('decks', []))).font = NF
        ws5.cell(row=row, column=4, value=info.get('owned', 0)).font = NF; ws5.cell(row=row, column=4).alignment = CTR
        surplus = info.get('surplus', 0)
        status = "OK" if surplus >= 0 else f"Need {abs(surplus)}"
        c = ws5.cell(row=row, column=5, value=status); c.alignment = CTR
        c.font = GF if surplus >= 0 else RF; c.fill = GRN if surplus >= 0 else RED
        row += 1
    ws5.column_dimensions['A'].width = 30; ws5.column_dimensions['B'].width = 30
    ws5.column_dimensions['C'].width = 45; ws5.column_dimensions['D'].width = 10; ws5.column_dimensions['E'].width = 14

    # --- CONSIDERING ---
    has_considering = {dn: cards for dn, cards in maybeboard.items() if cards}
    if has_considering:
        ws6 = wb.create_sheet("Considering (Maybeboard)"); ws6.sheet_properties.tabColor = "888888"
        hdr(ws6, 1, ["Card Name", "Deck", "In Collection?"])
        row = 2
        for dn in deck_names_raw:
            if dn not in has_considering: continue
            for card in sorted(maybeboard[dn].keys()):
                rc = resolve_cache.get(card, [card, 0])
                ws6.cell(row=row, column=1, value=card).font = NF
                ws6.cell(row=row, column=2, value=deck_display[dn]).font = NF
                status = "Yes" if rc[1] > 0 else "No"
                c = ws6.cell(row=row, column=3, value=status); c.alignment = CTR
                c.font = GF if rc[1] > 0 else RF; c.fill = GRN if rc[1] > 0 else RED
                row += 1
        ws6.column_dimensions['A'].width = 45; ws6.column_dimensions['B'].width = 42; ws6.column_dimensions['C'].width = 16

    # --- CONSIDERING DECKS TABS ---
    if considering_decks and assembly_order and available_pool is not None:
        con_display = {n: clean_deck_name(n + '.txt') for n in considering_decks}

        # Pre-compute per-deck overview stats (independent, before assembly claims)
        overview_rows = []
        for dn, cards in considering_decks.items():
            total = sum(cards.values())
            have = 0
            for card, needed in cards.items():
                cn = canonical_name(card)
                avail = available_pool.get(cn, 0)
                if cn not in available_pool:
                    avail = resolve_owned(card, collection)
                have += min(needed, avail)
            overview_rows.append((con_display[dn], considering_commanders.get(dn, ''),
                                  total, have, total - have,
                                  have / total if total else 0))
        overview_rows.sort(key=lambda x: -x[5])

        # -- Considering Overview --
        ws_co = wb.create_sheet("Considering Overview"); ws_co.sheet_properties.tabColor = "5B9BD5"
        ws_co['A1'] = "Considering Decks \u2013 Overview"
        ws_co['A1'].font = Font(bold=True, name='Arial', size=16, color='5B9BD5')
        ws_co.merge_cells('A1:F1')
        ws_co['A2'] = f"Cards available after existing {len(decks)} decks claim theirs"
        ws_co['A2'].font = Font(name='Arial', size=10, italic=True, color='888888')
        hdr(ws_co, 4, ["Deck", "Commander", "Cards", "Available", "Missing", "% Complete"])
        row = 5
        for display, cmd, total, have, miss, pct in overview_rows:
            ws_co.cell(row=row, column=1, value=display).font = BF
            ws_co.cell(row=row, column=2, value=cmd).font = Font(name='Arial', size=10, italic=True)
            ws_co.cell(row=row, column=3, value=total).font = NF; ws_co.cell(row=row, column=3).alignment = CTR
            ws_co.cell(row=row, column=4, value=have).font = NF; ws_co.cell(row=row, column=4).alignment = CTR
            c = ws_co.cell(row=row, column=5, value=miss); c.font = RF if miss > 0 else GF; c.alignment = CTR
            c.fill = RED if miss > 0 else GRN
            c = ws_co.cell(row=row, column=6, value=pct); c.number_format = '0.0%'; c.alignment = CTR
            if pct >= 0.9: c.font = GF; c.fill = GRN
            elif pct >= 0.7: c.font = Font(name='Arial', size=10, color='CC8800', bold=True); c.fill = YEL
            else: c.font = RF; c.fill = RED
            row += 1
        ws_co.column_dimensions['A'].width = 42; ws_co.column_dimensions['B'].width = 32
        ws_co.column_dimensions['C'].width = 10; ws_co.column_dimensions['D'].width = 12
        ws_co.column_dimensions['E'].width = 10; ws_co.column_dimensions['F'].width = 14
        ws_co.freeze_panes = 'A5'

        # -- Considering Shopping List --
        # Aggregate missing cards across all considering decks from assembly order
        shopping = defaultdict(lambda: {'need': 0, 'decks': []})
        for entry in assembly_order:
            for card, qty in entry['missing_cards'].items():
                cn = canonical_name(card)
                info = shopping[card]
                info['need'] += qty
                info['decks'].append(con_display[entry['name']])
                info['cn'] = cn
        if shopping:
            ws_cs = wb.create_sheet("Considering Shopping"); ws_cs.sheet_properties.tabColor = "C00000"
            hdr(ws_cs, 1, ["Card", "Copies Needed", "Owned", "Used by Existing", "Available", "Needed By"])
            row = 2
            for card in sorted(shopping.keys(), key=lambda c: (-shopping[c]['need'], c)):
                info = shopping[card]
                cn = info['cn']
                owned = resolve_owned(cn, collection)
                existing_demand = resolved_demand.get(cn, {}).get('total_demand', 0)
                avail = max(0, owned - existing_demand)
                ws_cs.cell(row=row, column=1, value=card).font = NF
                c = ws_cs.cell(row=row, column=2, value=info['need']); c.font = RF; c.alignment = CTR
                ws_cs.cell(row=row, column=3, value=owned).font = NF; ws_cs.cell(row=row, column=3).alignment = CTR
                ws_cs.cell(row=row, column=4, value=existing_demand).font = NF; ws_cs.cell(row=row, column=4).alignment = CTR
                ws_cs.cell(row=row, column=5, value=avail).font = NF; ws_cs.cell(row=row, column=5).alignment = CTR
                ws_cs.cell(row=row, column=6, value=", ".join(info['decks'])).font = NF
                row += 1
            ws_cs.column_dimensions['A'].width = 42; ws_cs.column_dimensions['B'].width = 16
            ws_cs.column_dimensions['C'].width = 10; ws_cs.column_dimensions['D'].width = 16
            ws_cs.column_dimensions['E'].width = 12; ws_cs.column_dimensions['F'].width = 80
            ws_cs.freeze_panes = 'A2'

        # -- Assembly Order --
        ws_ao = wb.create_sheet("Assembly Order"); ws_ao.sheet_properties.tabColor = "7030A0"
        ws_ao['A1'] = "Considering Decks \u2013 Recommended Build Order"
        ws_ao['A1'].font = Font(bold=True, name='Arial', size=16, color='7030A0')
        ws_ao.merge_cells('A1:G1')
        ws_ao['A2'] = "Greedy allocation: build the most-complete deck first, then re-evaluate"
        ws_ao['A2'].font = Font(name='Arial', size=10, italic=True, color='888888')
        hdr(ws_ao, 4, ["#", "Deck", "Commander", "Cards", "Available", "Missing", "% Complete"])
        row = 5
        for entry in assembly_order:
            ws_ao.cell(row=row, column=1, value=f"#{entry['order']}").font = BF; ws_ao.cell(row=row, column=1).alignment = CTR
            ws_ao.cell(row=row, column=2, value=con_display[entry['name']]).font = BF
            ws_ao.cell(row=row, column=3, value=entry['commander']).font = Font(name='Arial', size=10, italic=True)
            ws_ao.cell(row=row, column=4, value=entry['total']).font = NF; ws_ao.cell(row=row, column=4).alignment = CTR
            ws_ao.cell(row=row, column=5, value=entry['have']).font = NF; ws_ao.cell(row=row, column=5).alignment = CTR
            c = ws_ao.cell(row=row, column=6, value=entry['missing'])
            c.font = RF if entry['missing'] > 0 else GF; c.alignment = CTR
            c.fill = RED if entry['missing'] > 0 else GRN
            c = ws_ao.cell(row=row, column=7, value=entry['pct']); c.number_format = '0.0%'; c.alignment = CTR
            if entry['pct'] >= 0.9: c.font = GF; c.fill = GRN
            elif entry['pct'] >= 0.7: c.font = Font(name='Arial', size=10, color='CC8800', bold=True); c.fill = YEL
            else: c.font = RF; c.fill = RED
            row += 1
        ws_ao.column_dimensions['A'].width = 8; ws_ao.column_dimensions['B'].width = 42
        ws_ao.column_dimensions['C'].width = 32; ws_ao.column_dimensions['D'].width = 10
        ws_ao.column_dimensions['E'].width = 12; ws_ao.column_dimensions['F'].width = 10
        ws_ao.column_dimensions['G'].width = 14
        ws_ao.freeze_panes = 'A5'

    # --- PER-DECK TABS ---
    for dn in deck_names_raw:
        display = deck_display[dn]
        ws_d = wb.create_sheet(display[:31])
        cards = decks[dn]
        cmd = commanders.get(dn, "")
        hdr(ws_d, 1, ["Card Name", "Qty", "Owned", "Status", "Role"])

        def deck_sort(card):
            cn = canonical_name(card)
            info = resolved_demand.get(cn, {})
            return (0 if info.get('surplus', 0) < 0 else 1, card)

        row = 2
        for card in sorted(cards.keys(), key=deck_sort):
            needed = cards[card]
            cn = canonical_name(card)
            rc = resolve_cache.get(card, [cn, 0])
            owned = rc[1]
            info = resolved_demand.get(cn, {'surplus': 0})
            surplus = info.get('surplus', 0)
            if surplus >= 0: status = "OK"
            elif owned >= needed: status = f"Shared conflict ({surplus})"
            else: status = "Not in collection"
            role = "Commander" if card == cmd else ""
            alias_note = f" (={ALIASES[card]})" if card in ALIASES else ""
            ws_d.cell(row=row, column=1, value=card + alias_note).font = BF if card == cmd else NF
            ws_d.cell(row=row, column=2, value=needed).font = NF; ws_d.cell(row=row, column=2).alignment = CTR
            ws_d.cell(row=row, column=3, value=owned).font = NF; ws_d.cell(row=row, column=3).alignment = CTR
            c = ws_d.cell(row=row, column=4, value=status)
            if "Not in" in status: c.font = RF; c.fill = RED
            elif "Shared" in status: c.font = Font(name='Arial', size=10, color='CC8800', bold=True); c.fill = YEL
            else: c.font = GF; c.fill = GRN
            if role: ws_d.cell(row=row, column=5, value=role).font = Font(name='Arial', size=10, color='4472C4', bold=True)
            row += 1
        ws_d.column_dimensions['A'].width = 50; ws_d.column_dimensions['B'].width = 8
        ws_d.column_dimensions['C'].width = 10; ws_d.column_dimensions['D'].width = 30; ws_d.column_dimensions['E'].width = 14
        ws_d.freeze_panes = 'A2'

    if output_path and hasattr(wb, 'save'):
        wb.save(output_path)
    return missing, shared, wb


# ============================================================
# GOOGLE SHEETS UPLOAD
# ============================================================
SHEET_ID_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.deck_safe_sheet_id')


def _load_cached_sheet_id():
    if os.path.exists(SHEET_ID_FILE):
        return open(SHEET_ID_FILE).read().strip()
    return None


def _save_cached_sheet_id(sheet_id):
    with open(SHEET_ID_FILE, 'w') as f:
        f.write(sheet_id)


def upload_to_google_sheets(gwb, sheet_name):
    import gspread

    gc = gspread.oauth()

    spreadsheet = None
    sheet_id = _load_cached_sheet_id()

    if sheet_id:
        try:
            spreadsheet = gc.open_by_key(sheet_id)
        except gspread.exceptions.SpreadsheetNotFound:
            spreadsheet = None

    if not spreadsheet:
        try:
            spreadsheet = gc.open(sheet_name)
        except gspread.exceptions.SpreadsheetNotFound:
            spreadsheet = gc.create(sheet_name)
            print(f"  Created new Google Sheet: {sheet_name}")

    _save_cached_sheet_id(spreadsheet.id)

    print(f"  Uploading to Google Sheets...")
    gwb.flush_to_google(spreadsheet)
    print(f"  Google Sheets: {spreadsheet.url}")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='Build a Deck-Safe Collection spreadsheet')
    parser.add_argument('csv', help='Path to Moxfield CSV haves export')
    parser.add_argument('decks', nargs='*', help='Deck .txt files')
    parser.add_argument('--deck-dir', '-d', help='Directory containing deck .txt files')
    parser.add_argument('-o', '--output', default=None, help='Output xlsx path (local file)')
    parser.add_argument('--no-google', action='store_true', help='Skip Google Sheets upload')
    parser.add_argument('--sheet-name', default='Deck-Safe Collection', help='Google Sheets document name')
    parser.add_argument('--no-proxy', action='store_true', help='Skip proxy file generation')
    parser.add_argument('--proxy-dir', default=None, help='Directory for proxy .txt output')
    args = parser.parse_args()

    # Gather deck files
    deck_files = list(args.decks)
    if args.deck_dir:
        for f in sorted(os.listdir(args.deck_dir)):
            if f.endswith('.txt'):
                deck_files.append(os.path.join(args.deck_dir, f))

    # Discover considering decks
    considering_dir = None
    considering_files = []
    if args.deck_dir:
        considering_dir = os.path.join(args.deck_dir, 'considering')
        if os.path.isdir(considering_dir):
            considering_files = sorted(
                os.path.join(considering_dir, f)
                for f in os.listdir(considering_dir)
                if f.endswith('.txt')
            )

    if not deck_files:
        print("ERROR: No deck files provided. Use positional args or --deck-dir")
        sys.exit(1)

    print(f"Collection: {args.csv}")
    print(f"Decks: {len(deck_files)} files")
    if considering_files:
        print(f"Considering: {len(considering_files)} files from {considering_dir}")
    if args.output:
        print(f"Output: {args.output}")
    if not args.no_google:
        print(f"Google Sheet: {args.sheet_name}")
    print()

    # Parse collection
    collection_raw, collection = parse_collection(args.csv)
    print(f"Collection: {len(collection_raw)} unique cards, {sum(collection_raw.values())} total copies")

    # Parse decks
    decks = {}
    maybeboard_all = {}
    commanders = {}
    for filepath in deck_files:
        filename = os.path.basename(filepath)
        deck_name = filename.replace('.txt', '')
        actual_deck, sb_cards, commander = parse_deck(filepath)
        decks[deck_name] = actual_deck
        maybeboard_all[deck_name] = sb_cards
        if commander:
            commanders[deck_name] = commander
        total = sum(actual_deck.values())
        sb_total = sum(sb_cards.values())
        print(f"  {clean_deck_name(filename)}: {total} cards + {sb_total} considering [{commander or '?'}]")

    # Parse considering decks
    considering_decks = {}
    considering_commanders = {}
    for filepath in considering_files:
        filename = os.path.basename(filepath)
        deck_name = filename.replace('.txt', '')
        actual_deck, _, commander = parse_deck(filepath)
        considering_decks[deck_name] = actual_deck
        if commander:
            considering_commanders[deck_name] = commander
        total = sum(actual_deck.values())
        print(f"  [Considering] {clean_deck_name(filename)}: {total} cards [{commander or '?'}]")

    # Calculate demand
    resolved_demand = defaultdict(lambda: {'total_demand': 0, 'owned': 0, 'decks': [], 'display_names': []})
    for dn, cards in decks.items():
        for card, cnt in cards.items():
            cn = canonical_name(card)
            info = resolved_demand[cn]
            info['total_demand'] += cnt
            info['decks'].append(dn)
            if card not in info['display_names']:
                info['display_names'].append(card)

    for cn, info in resolved_demand.items():
        info['owned'] = resolve_owned(cn, collection)
        info['surplus'] = info['owned'] - info['total_demand']

    resolve_cache = {}
    all_cards = set(c for d in decks.values() for c in d) | set(c for d in maybeboard_all.values() for c in d)
    all_cards |= set(c for d in considering_decks.values() for c in d)
    for card in all_cards:
        cn = canonical_name(card)
        resolve_cache[card] = [cn, resolve_owned(cn, collection)]

    # Compute considering decks analysis
    available_pool = compute_available_pool(collection, resolved_demand)
    assembly_order = []
    if considering_decks:
        assembly_order = compute_assembly_order(
            considering_decks, considering_commanders, available_pool, collection)

    # Build spreadsheet
    deck_names_raw = sorted(decks.keys())
    deck_display = {n: clean_deck_name(n + '.txt') for n in deck_names_raw}
    build_args = (collection_raw, collection, decks, maybeboard_all, commanders,
                  deck_names_raw, deck_display, dict(resolved_demand), resolve_cache,
                  os.path.basename(args.csv))
    build_kwargs = {
        'considering_decks': considering_decks,
        'considering_commanders': considering_commanders,
        'assembly_order': assembly_order,
        'available_pool': available_pool,
    }

    # Local .xlsx output
    if args.output:
        missing, shared, _ = build_spreadsheet(*build_args, output_path=args.output, **build_kwargs)
        print(f"\n  Saved to {args.output}")
    else:
        missing = sum(1 for v in resolved_demand.values() if v['surplus'] < 0)
        shared = sum(1 for v in resolved_demand.values() if len(v['decks']) > 1)

    # Google Sheets upload
    if not args.no_google:
        gwb = GWorkbook()
        build_spreadsheet(*build_args, output_path=None, workbook=gwb, **build_kwargs)
        upload_to_google_sheets(gwb, args.sheet_name)

    # Write proxy files
    if considering_decks and not args.no_proxy:
        proxy_dir = args.proxy_dir or (os.path.join(args.deck_dir, 'proxy') if args.deck_dir else 'proxy')
        written = write_proxy_files(assembly_order, proxy_dir)
        if written:
            print(f"\n  Proxy files written to {proxy_dir}/:")
            for p in written:
                print(f"    {os.path.basename(p)}")

    print(f"\n{'='*50}")
    print(f"  {len(decks)} decks, {len(resolved_demand)} unique cards")
    print(f"  {missing} missing, {shared} shared")
    if assembly_order:
        print(f"  {len(considering_decks)} considering decks analyzed")
    print(f"{'='*50}")


if __name__ == '__main__':
    main()
